import pandas as pd
import openpyxl
from openpyxl.styles import Font, Alignment
import logging
import os
from datetime import datetime
from collections import defaultdict
import math
import time
import csv

# Configure logging
logging.basicConfig(
    filename='errors.txt',
    level=logging.ERROR,
    format='%(asctime)s - %(levelname)s - %(message)s'
)

class ExamSeatingArrangement:
    def __init__(self, buffer=1, mode='dense'):
        """
        Initialize the exam seating arrangement system.
        
        Args:
            buffer (int): Number of buffer seats per classroom
            mode (str): 'dense' or 'sparse' seating mode
        """
        self.start_time = None
        self.last_log_time = None
        self.buffer = buffer
        self.mode = mode
        self.validate_mode()
        
        # Data storage
        self.exam_schedule = None
        self.course_roll_mapping = None
        self.roll_name_mapping = None
        self.classroom_master = None
        
        # File-based storage for seating arrangements
        self.seating_arrangement_file = "temp_allocations.csv"
        self.initialize_temp_file()
        
        # Output containers
        self.seats_left = []
        self.clashes = []
    
    def initialize_temp_file(self):
        """Initialize or clear the temporary seating arrangement file."""
        if os.path.exists(self.seating_arrangement_file):
            os.remove(self.seating_arrangement_file)
        # Create empty file with header
        with open(self.seating_arrangement_file, 'w', newline='') as f:
            writer = csv.DictWriter(f, fieldnames=[
                'date', 'day', 'session', 'course_code', 'room', 'block',
                'allocated_student_count', 'roll_list', 'exam_capacity',
                'adjusted_capacity', 'floor'
            ])
            writer.writeheader()
    
    def validate_mode(self):
        """Validate the seating mode is either 'dense' or 'sparse'."""
        if self.mode not in ['dense', 'sparse']:
            raise ValueError("Mode must be either 'dense' or 'sparse'")
    
    def load_data(self, file_path):
        """
        Load data from the input Excel file with multiple sheets.
        
        Args:
            file_path (str): Path to the input Excel file
        """
        try:
            xls = pd.ExcelFile(file_path)
            
            # Load each sheet
            self.exam_schedule = pd.read_excel(xls, sheet_name='in_timetable')
            self.course_roll_mapping = pd.read_excel(xls, sheet_name='in_course_roll_mapping')
            self.roll_name_mapping = pd.read_excel(xls, sheet_name='in_roll_name_mapping')
            self.classroom_master = pd.read_excel(xls, sheet_name='in_room_capacity')
            
            # Validate data before processing
            self.validate_input_data()
            
            # Clean and standardize column names
            self.standardize_data()
            
            # Pre-process classroom data to extract floor information
            self.extract_floor_info()
            
            # Initialize adjusted capacities
            self.classroom_master['adjusted_capacity'] = self.classroom_master['exam capacity'] - self.buffer
            
        except Exception as e:
            logging.error(f"Error loading data from {file_path}: {str(e)}", exc_info=True)
            raise
    
    def validate_input_data(self):
        """Validate that all required columns exist in input data."""
        required_columns = {
            'in_timetable': ['date', 'day', 'morning', 'evening'],
            'in_course_roll_mapping': ['course_code', 'rollno'],
            'in_roll_name_mapping': ['roll', 'name'],
            'in_room_capacity': ['room no.', 'exam capacity', 'block']
        }
        
        for sheet, columns in required_columns.items():
            try:
                df = getattr(self, sheet.replace('in_', '')) if hasattr(self, sheet.replace('in_', '')) else None
                if df is not None:
                    for col in columns:
                        if col not in df.columns:
                            raise ValueError(f"Missing required column '{col}' in sheet {sheet}")
            except Exception as e:
                logging.error(f"Validation error in {sheet}: {str(e)}")
                raise
    
    def extract_floor_info(self):
        """Extract floor information from room numbers."""
        if self.classroom_master is not None and 'room no.' in self.classroom_master.columns:
            self.classroom_master['floor'] = self.classroom_master['room no.'].apply(
                lambda x: str(x)[0] if pd.notna(x) and str(x)[0].isdigit() else '0'
            )
    
    def standardize_data(self):
        """Standardize column names and data formats across dataframes."""
        try:
            # Standardize column names
            self.exam_schedule.columns = [col.strip().lower() for col in self.exam_schedule.columns]
            self.course_roll_mapping.columns = [col.strip().lower() for col in self.course_roll_mapping.columns]
            self.roll_name_mapping.columns = [col.strip().lower() for col in self.roll_name_mapping.columns]
            self.classroom_master.columns = [col.strip().lower() for col in self.classroom_master.columns]

            # Parse and expand the exam schedule
            self.exam_schedule = self.parse_exam_schedule(self.exam_schedule)

            # Ensure roll_no is string type
            if 'rollno' in self.course_roll_mapping.columns:
                self.course_roll_mapping['rollno'] = self.course_roll_mapping['rollno'].astype(str)
            if 'roll' in self.roll_name_mapping.columns:
                self.roll_name_mapping['roll'] = self.roll_name_mapping['roll'].astype(str)

            # Fill missing names with "Unknown Name"
            if 'name' in self.roll_name_mapping.columns:
                self.roll_name_mapping['name'] = self.roll_name_mapping['name'].fillna('Unknown Name')
            
        except Exception as e:
            logging.error(f"Error standardizing data: {str(e)}", exc_info=True)
            raise
    
    def parse_exam_schedule(self, schedule_df):
        """Parse the exam schedule with combined subject codes."""
        parsed_rows = []
        
        for _, row in schedule_df.iterrows():
            date = row['date']
            # Convert to string if it's a datetime object
            if isinstance(date, pd.Timestamp):
                date = date.strftime('%Y-%m-%d')  # or your preferred string format
            elif not isinstance(date, str):
                date = str(date)
                
            day = row['day']
            
            # Rest of the method remains the same
            if pd.notna(row['morning']) and str(row['morning']).strip().lower() != 'no exam':
                subjects = [s.strip() for s in str(row['morning']).split(';') if s.strip()]
                for subject in subjects:
                    parsed_rows.append({
                        'date': date,
                        'day': day,
                        'session': 'morning',
                        'subject_code': subject
                    })
            
            if pd.notna(row['evening']) and str(row['evening']).strip().lower() != 'no exam':
                subjects = [s.strip() for s in str(row['evening']).split(';') if s.strip()]
                for subject in subjects:
                    parsed_rows.append({
                        'date': date,
                        'day': day,
                        'session': 'evening',
                        'subject_code': subject
                    })
        
        return pd.DataFrame(parsed_rows)
    
    def detect_clashes(self):
        """Detect exam clashes for students."""
        try:
            # Group exams by date and session
            grouped_exams = self.exam_schedule.groupby(['date', 'day', 'session'])
            
            for (date, day, session), exams in grouped_exams:
                subject_codes = exams['subject_code'].unique()
                
                # Get all unique pairs of subjects
                from itertools import combinations
                for subj1, subj2 in combinations(subject_codes, 2):
                    # Get roll numbers for each subject
                    rolls1 = set(self.get_rolls_for_subject(subj1))
                    rolls2 = set(self.get_rolls_for_subject(subj2))
                    
                    # Find intersection
                    common_rolls = rolls1 & rolls2
                    
                    if common_rolls:
                        clash_info = {
                            'date': date,
                            'day': day,
                            'session': session,
                            'subjects': [subj1, subj2],
                            'roll_numbers': list(common_rolls)
                        }
                        self.clashes.append(clash_info)
                        
                        # Log the clash
                        for roll in common_rolls:
                            logging.error(
                                f"Roll No {roll} has multiple exams on {date}, {session} in {subj1} and {subj2}"
                            )
                            print(
                                f"Roll No {roll} has multiple exams on {date}, {session} in {subj1} and {subj2}"
                            )
            
            return len(self.clashes) == 0  # Returns True if no clashes
        
        except Exception as e:
            logging.error(f"Error detecting clashes: {str(e)}", exc_info=True)
            raise
    
    def get_rolls_for_subject(self, subject_code):
        """Get all roll numbers for a given subject code."""
        try:
            rolls = self.course_roll_mapping[
                self.course_roll_mapping['course_code'] == subject_code
            ]['rollno'].unique()
            return list(rolls)
        except Exception as e:
            logging.error(f"Error getting rolls for subject {subject_code}: {str(e)}", exc_info=True)
            return []
    
    def get_student_name(self, roll_no):
        """Get student name for a given roll number."""
        try:
            name = self.roll_name_mapping[
                self.roll_name_mapping['roll'] == str(roll_no)
            ]['name'].values[0]
            return name
        except (IndexError, KeyError):
            logging.warning(f"Name not found for roll number {roll_no}")
            return "Unknown Name"
    
    def allocate_seats(self):
        """Allocate seats for all exams based on the given constraints."""
        self.start_time = time.time()
        self.last_log_time = self.start_time
        try:
            logging.info("Starting seat allocation process")
            print("\nStarting seat allocation process...")
            
            # Group exams by date and session
            grouped_exams = self.exam_schedule.groupby(['date', 'day', 'session'])
            total_groups = len(grouped_exams)
            current_group = 0
            
            for (date, day, session), exams in grouped_exams:
                current_group += 1
                group_start_time = time.time()
                logging.info(f"Processing group {current_group}/{total_groups}: {date} {session}")
                print(f"\nProcessing exams on {date} {session} ({current_group}/{total_groups})...")
                
                subject_codes = exams['subject_code'].unique()
                total_subjects = len(subject_codes)
                processed_subjects = 0
                
                # Log subject count
                logging.info(f"Found {total_subjects} subjects for {date} {session}")
                
                # Sort subjects by number of students (descending)
                subjects_with_counts = []
                for subj in subject_codes:
                    rolls = self.get_rolls_for_subject(subj)
                    subjects_with_counts.append((subj, len(rolls)))
                
                subjects_with_counts.sort(key=lambda x: x[1], reverse=True)
                
                # Prepare available rooms (grouped by floor)
                available_rooms_by_floor = defaultdict(list)
                
                for _, room in self.classroom_master.sort_values(
                    'exam capacity', ascending=False
                ).iterrows():
                    room_dict = room.to_dict()
                    available_rooms_by_floor[room_dict['floor']].append(room_dict)
                
                # Log room information
                total_rooms = sum(len(rooms) for rooms in available_rooms_by_floor.values())
                logging.info(f"Available rooms: {total_rooms} across {len(available_rooms_by_floor)} floors")
                
                # Allocate rooms for each subject
                for subj, student_count in subjects_with_counts:
                    processed_subjects += 1
                    subject_start_time = time.time()
                    
                    logging.info(f"Allocating {student_count} students for {subj} "
                                f"({processed_subjects}/{total_subjects})")
                    print(f"  Allocating {subj} ({student_count} students)...", end=' ')
                    
                    self.allocate_subject_rooms(
                        subj, student_count, available_rooms_by_floor, date, day, session
                    )
                    
                    # Log progress
                    subject_time = time.time() - subject_start_time
                    logging.info(f"Completed {subj} in {subject_time:.2f} seconds")
                    print(f"done in {subject_time:.2f}s")
                    
                    # Periodic status update
                    self._log_progress(current_group, total_groups, processed_subjects, total_subjects)
                
                group_time = time.time() - group_start_time
                logging.info(f"Completed group {date} {session} in {group_time:.2f} seconds")
            
            # Final timing
            total_time = time.time() - self.start_time
            logging.info(f"Seat allocation completed in {total_time:.2f} seconds")
            print(f"\nSeat allocation completed in {total_time:.2f} seconds")
            
            # Prepare seats left data
            logging.info("Preparing seats left data")
            self.prepare_seats_left_data()
            
        except Exception as e:
            logging.error(f"Error allocating seats: {str(e)}", exc_info=True)
            raise
    
    def _log_progress(self, current_group, total_groups, processed_subjects, total_subjects):
        """Log periodic progress updates."""
        now = time.time()
        if now - self.last_log_time > 30:  # Log every 30 seconds
            elapsed = now - self.start_time
            progress = (current_group / total_groups) * 100
            logging.info(
                f"Progress: {progress:.1f}% - "
                f"Group {current_group}/{total_groups} - "
                f"Subject {processed_subjects}/{total_subjects} - "
                f"Elapsed: {elapsed:.1f}s"
            )
            print(
                f"\nProgress update: {progress:.1f}% complete - "
                f"Elapsed time: {elapsed:.1f}s"
            )
            self.last_log_time = now

    def select_best_floor(self, available_rooms_by_floor, remaining_students):
        """Select the best floor to allocate rooms from, preferring floors with existing allocations."""
        # First try to find a floor with existing allocations that can accommodate
        for floor, rooms in available_rooms_by_floor.items():
            if rooms and sum(r['adjusted_capacity'] for r in rooms) >= remaining_students:
                return floor
        
        # If no suitable floor with existing allocations, find the floor with largest capacity
        best_floor = None
        max_capacity = 0
        for floor, rooms in available_rooms_by_floor.items():
            floor_capacity = sum(r['adjusted_capacity'] for r in rooms)
            if floor_capacity > max_capacity:
                max_capacity = floor_capacity
                best_floor = floor
        return best_floor
    
    def select_best_room(self, rooms, remaining_students):
        """Select the most appropriate room based on mode and capacity."""
        try:
            if self.mode == 'dense':
                # Find smallest room that can fit all remaining students
                suitable_rooms = [r for r in rooms if r['adjusted_capacity'] >= remaining_students]
                if suitable_rooms:
                    return suitable_rooms[-1]  # Take the smallest suitable room
                else:
                    return rooms[0]  # Take the largest available room
            else:  # sparse mode
                return rooms[0]  # Take the largest available room
        except Exception as e:
            logging.error(f"Error selecting best room: {str(e)}", exc_info=True)
            return None
    
    def allocate_subject_rooms(self, subject_code, student_count, available_rooms_by_floor, date, day, session):
        """Allocate rooms for a specific subject with proper capacity enforcement."""
        try:
            remaining_students = student_count
            rolls = self.get_rolls_for_subject(subject_code)
            iteration = 0
            max_iterations = 100  # Prevent infinite loops
            
            while remaining_students > 0 and iteration < max_iterations:
                iteration += 1
                
                # Debug logging
                if iteration % 10 == 0:
                    logging.info(f"Iteration {iteration} for {subject_code}: {remaining_students} students remaining")
                    print(f"    Iteration {iteration}: {remaining_students} students remaining")
                
                # Select best floor to allocate from
                floor = self.select_best_floor(available_rooms_by_floor, remaining_students)
                if not floor or not available_rooms_by_floor[floor]:
                    logging.warning(f"No available rooms on floor {floor} for {subject_code}")
                    break
                    
                # Select best room from this floor
                room = self.select_best_room(available_rooms_by_floor[floor], remaining_students)
                if room is None:
                    logging.warning(f"No suitable room found for {subject_code}")
                    break
                    
                # Calculate how many students we can allocate
                students_to_allocate = min(
                    remaining_students,
                    math.floor(room['adjusted_capacity'] * 0.5) if self.mode == 'sparse' 
                    else room['adjusted_capacity']
                )
                
                # Ensure we allocate at least 1 student
                if students_to_allocate <= 0:
                    logging.warning(f"Room {room['room no.']} has no capacity left")
                    available_rooms_by_floor[floor].remove(room)
                    continue
                    
                # Get ACTUAL roll numbers to allocate
                allocated_rolls = rolls[:students_to_allocate]
                actual_count = len(allocated_rolls)
                
                # Write allocation to file
                self.write_allocation_to_file({
                    'date': date,
                    'day': day,
                    'session': session,
                    'course_code': subject_code,
                    'room': room['room no.'],
                    'block': room['block'],
                    'allocated_student_count': actual_count,
                    'roll_list': ';'.join(allocated_rolls),
                    'exam_capacity': room['exam capacity'],
                    'adjusted_capacity': room['adjusted_capacity'],
                    'floor': room['floor']
                })
                
                # Update remaining students and room availability
                remaining_students -= actual_count
                rolls = rolls[actual_count:]
                
                # Update room capacity
                room['adjusted_capacity'] -= actual_count
                if room['adjusted_capacity'] <= 0:
                    available_rooms_by_floor[floor].remove(room)
                else:
                    # Re-sort rooms on this floor
                    available_rooms_by_floor[floor].sort(
                        key=lambda x: x['adjusted_capacity'], reverse=True
                    )
                
            if remaining_students > 0:
                error_msg = f"Cannot allocate all students for {subject_code} on {date} {session}. {remaining_students} students left."
                logging.error(error_msg)
                print(error_msg)
                
            if iteration >= max_iterations:
                error_msg = f"Reached maximum iterations ({max_iterations}) for {subject_code}"
                logging.error(error_msg)
                print(error_msg)
                
        except Exception as e:
            logging.error(f"Error allocating rooms for {subject_code}: {str(e)}", exc_info=True)
            raise

    def write_allocation_to_file(self, allocation):
        """Write allocation data to the temporary file."""
        with open(self.seating_arrangement_file, 'a', newline='') as f:
            writer = csv.DictWriter(f, fieldnames=allocation.keys())
            writer.writerow(allocation)
    
    def get_seating_arrangement(self):
        """Read seating arrangement from temporary file."""
        try:
            with open(self.seating_arrangement_file, 'r') as f:
                reader = csv.DictReader(f)
                return list(reader)
        except FileNotFoundError:
            return []
    
    def prepare_seats_left_data(self):
        """Calculate remaining capacity with proper validation."""
        try:
            seating_data = self.get_seating_arrangement()
            if not seating_data:
                return
                
            # Create DataFrame with ACTUAL student counts
            room_usage = pd.DataFrame(seating_data)
            room_usage['actual_count'] = room_usage['roll_list'].apply(
                lambda x: len(x.split(';')) if pd.notna(x) else 0
            )
            
            # Convert numeric columns
            room_usage['exam_capacity'] = pd.to_numeric(room_usage['exam_capacity'])
            room_usage['actual_count'] = pd.to_numeric(room_usage['actual_count'])
            
            # Group by room and calculate totals
            room_allocation = room_usage.groupby(['room', 'block', 'exam_capacity', 'floor']).agg({
                'actual_count': 'sum'
            }).reset_index()
            
            # Calculate vacant seats properly
            room_allocation['vacant'] = room_allocation['exam_capacity'] - room_allocation['actual_count']
            room_allocation['vacant'] = room_allocation['vacant'].clip(lower=0)  # No negative vacancies
            
            # Convert to output format
            self.seats_left = room_allocation[[
                'room', 'exam_capacity', 'block', 'floor', 'actual_count', 'vacant'
            ]].rename(columns={
                'room': 'room_no',
                'actual_count': 'allocated'
            }).to_dict('records')
            
        except Exception as e:
            logging.error(f"Error preparing seats left data: {str(e)}", exc_info=True)
            raise
    
    def generate_output_files(self, output_folder='output'):
        """Generate all output files and folder structure."""
        try:
            # Create main output folder
            os.makedirs(output_folder, exist_ok=True)
            
            # Generate overall files
            self.generate_overall_files(output_folder)
            
            # Generate per-date files
            self.generate_per_date_files(output_folder)
            
            # Clean up temporary files
            self.cleanup_temp_files()
            
        except Exception as e:
            logging.error(f"Error generating output files: {str(e)}", exc_info=True)
            raise
    
    def cleanup_temp_files(self):
        """Remove temporary files."""
        if os.path.exists(self.seating_arrangement_file):
            os.remove(self.seating_arrangement_file)
    
    def generate_overall_files(self, output_folder):
        """Generate the overall seating arrangement and seats left files."""
        try:
            # Overall seating arrangement
            seating_data = self.get_seating_arrangement()
            overall_seating = pd.DataFrame(seating_data)
            overall_seating = overall_seating[[
                'date', 'day', 'course_code', 'room', 'block', 'floor', 'allocated_student_count'
            ]]
            
            # Save to Excel with formatting
            output_path = os.path.join(output_folder, 'op_overall_seating_arrangement.xlsx')
            with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
                overall_seating.to_excel(writer, index=False, sheet_name='Seating Plan')
                
                # Format the worksheet
                workbook = writer.book
                worksheet = writer.sheets['Seating Plan']
                
                # Add title
                worksheet.insert_rows(1)
                worksheet.cell(row=1, column=1, value="Seating Plan")
                worksheet.cell(row=1, column=1).font = Font(bold=True, size=14)
                worksheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=7)
            
            # Seats left file
            seats_left_df = pd.DataFrame(self.seats_left)
            seats_left_df.to_excel(
                os.path.join(output_folder, 'op_seats_left.xlsx'),
                index=False
            )
            
        except Exception as e:
            logging.error(f"Error generating overall files: {str(e)}", exc_info=True)
            raise
    
    def generate_per_date_files(self, output_folder):
        """Generate the per-date folder structure and files."""
        try:
            seating_data = self.get_seating_arrangement()
            if not seating_data:
                return
                
            seating_df = pd.DataFrame(seating_data)
            grouped = seating_df.groupby(['date', 'session'])
            
            for (date, session), group in grouped:
                # Ensure date is string
                if not isinstance(date, str):
                    date = str(date)
                    
                # Format date for folder name (assuming input is YYYY-MM-DD)
                try:
                    date_parts = date.split('-')
                    if len(date_parts) == 3:
                        date_str = f"{date_parts[2]}_{date_parts[1]}_{date_parts[0]}"  # DD_MM_YYYY
                    else:
                        date_str = date.replace('-', '_')  # fallback
                except:
                    date_str = date.replace('-', '_')  # fallback
                    
                session_str = 'morning' if session == 'morning' else 'evening'
                
                # Create folders
                date_folder = os.path.join(output_folder, date_str)
                session_folder = os.path.join(date_folder, session_str)
                os.makedirs(session_folder, exist_ok=True)
                
                # Rest of the method remains the same...
                
                # Group by subject and room
                subject_room_group = group.groupby(['course_code', 'room'])
                
                for (subject_code, room), subject_group in subject_room_group:
                    # Generate attendance sheet
                    self.generate_attendance_sheet(
                        subject_code, room, subject_group, session_folder, date, session
                    )
                    
                    # Generate seating arrangement file
                    self.generate_seating_arrangement_file(
                        subject_code, room, subject_group, session_folder
                    )
                    
        except Exception as e:
            logging.error(f"Error generating per-date files: {str(e)}", exc_info=True)
            raise
    
    def generate_attendance_sheet(self, subject_code, room, subject_group, session_folder, date, session):
        """Generate attendance sheet for a subject in a specific room."""
        try:
            # Get roll numbers and names
            roll_list = subject_group['roll_list'].values[0].split(';')
            student_data = []
            
            for roll in roll_list:
                name = self.get_student_name(roll)
                if name == 'Unknown Name':
                    name = f'Unknown Name (Roll: {roll})'
                student_data.append({
                    'Roll No': roll,
                    'Student Name': name,
                    'Signature': ''
                })
            
            # Create DataFrame
            attendance_df = pd.DataFrame(student_data)
            
            # Save to Excel
            filename = f"attendance_sheet_{subject_code}_{room}.xlsx"
            filepath = os.path.join(session_folder, filename)
            
            with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
                attendance_df.to_excel(writer, index=False, sheet_name='Attendance')
                
                # Format the worksheet
                workbook = writer.book
                worksheet = writer.sheets['Attendance']
                
                # Add title with course info
                worksheet.insert_rows(1)
                title = f"{subject_code} - Room: {room} - Date: {date} - Session: {session}"
                worksheet.cell(row=1, column=1, value=title)
                worksheet.cell(row=1, column=1).font = Font(bold=True, size=14)
                worksheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=3)
                
                # Add TA and invigilator sections
                start_row = len(attendance_df) + 4
                for i in range(1, 6):
                    worksheet.cell(row=start_row + i - 1, column=1, value=f"TA{i}:")
                
                for i in range(1, 6):
                    worksheet.cell(row=start_row + 5 + i - 1, column=1, value=f"Invigilator {i}:")
            
        except Exception as e:
            logging.error(f"Error generating attendance sheet for {subject_code} in {room}: {str(e)}")
            raise
    
    def generate_seating_arrangement_file(self, subject_code, room, subject_group, session_folder):
        """Generate seating arrangement file for a subject in a specific room."""
        try:
            # Get the first row (all rows should have same date, day, etc.)
            row = subject_group.iloc[0]
            
            # Create DataFrame with the arrangement info
            arrangement_data = [{
                'date': row['date'],
                'day': row['day'],
                'course_code': row['course_code'],
                'room': row['room'],
                'allocated_student_count': row['allocated_student_count'],
                'roll_list': row['roll_list']
            }]
            arrangement_df = pd.DataFrame(arrangement_data)
            
            # Save to Excel with formatting
            filename = f"seating_arrangement_{subject_code}_{room}.xlsx"
            filepath = os.path.join(session_folder, filename)
            
            with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
                arrangement_df.to_excel(writer, index=False, sheet_name='Seating Plan')
                
                # Format the worksheet
                workbook = writer.book
                worksheet = writer.sheets['Seating Plan']
                
                # Add title
                worksheet.insert_rows(1)
                worksheet.cell(row=1, column=1, value="Seating Plan")
                worksheet.cell(row=1, column=1).font = Font(bold=True, size=14)
                worksheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=6)
            
        except Exception as e:
            logging.error(f"Error generating seating arrangement for {subject_code} in {room}: {str(e)}")
            raise
    
    def generate_seating_arrangement_file(self, subject_code, room, subject_group, session_folder):
        """Generate seating arrangement file for a subject in a specific room."""
        try:
            # Get the first row (all rows should have same date, day, etc.)
            row = subject_group.iloc[0]
            
            # Create DataFrame with the arrangement info
            arrangement_data = [{
                'date': row['date'],
                'day': row['day'],
                'course_code': row['course_code'],
                'room': row['room'],
                'block': row['block'],
                'floor': row['floor'],
                'allocated_student_count': row['allocated_student_count'],
                'roll_list': row['roll_list']
            }]
            arrangement_df = pd.DataFrame(arrangement_data)
            
            # Save to Excel with formatting
            filename = f"seating_arrangement_{subject_code}_{room}.xlsx"
            filepath = os.path.join(session_folder, filename)
            
            with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
                arrangement_df.to_excel(writer, index=False, sheet_name='Seating Plan')
                
                # Format the worksheet
                workbook = writer.book
                worksheet = writer.sheets['Seating Plan']
                
                # Add title
                worksheet.insert_rows(1)
                worksheet.cell(row=1, column=1, value="Seating Plan")
                worksheet.cell(row=1, column=1).font = Font(bold=True, size=14)
                worksheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=7)
            
        except Exception as e:
            logging.error(f"Error generating seating arrangement for {subject_code} in {room}: {str(e)}", exc_info=True)
            raise


def main():
    try:
        # User inputs
        buffer_seats = int(input("Enter number of buffer seats per classroom: "))
        mode = input("Enter seating mode (dense/sparse): ").lower()
        
        # Initialize system
        seating_system = ExamSeatingArrangement(buffer=buffer_seats, mode=mode)
        
        # Load data
        input_file = 'sorted_output_file.xlsx'
        seating_system.load_data(input_file)
        
        # Detect clashes
        print("\nChecking for exam clashes...")
        no_clashes = seating_system.detect_clashes()
        
        if no_clashes:
            print("No exam clashes detected.")
        else:
            print("Exam clashes detected. Check errors.txt for details.")
        
        # Allocate seats
        print("\nAllocating seats...")
        seating_system.allocate_seats()
        
        # Generate output files
        print("\nGenerating output files...")
        seating_system.generate_output_files()
        
        print("\nProcess completed successfully!")
        
    except Exception as e:
        logging.error(f"Error in main execution: {str(e)}", exc_info=True)
        print(f"An error occurred: {str(e)}. Check errors.txt for details.")


if __name__ == "__main__":
    main()